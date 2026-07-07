#!/usr/bin/env python3
"""Create empty output templates for paper-topic-builder."""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

SCHEMAS = {
    "literature_matrix.xlsx": [
        "Paper_ID", "Zotero_Key", "Title", "Authors", "Year", "Journal", "DOI",
        "Abstract", "Keywords", "Zotero_Tags", "Collection_Path", "Attachment_Status",
        "Analysis_Basis", "Research_Type", "Type_Evidence", "Type_Confidence",
        "Research_Topic", "Research_Background", "Research_Object", "Research_Context",
        "Country_Region", "Theoretical_Foundation", "Core_Research_Question",
        "Research_Method", "Sample_Source", "Sample_Size", "Data_Type", "Core_Findings",
        "Theoretical_Contribution", "Practical_Contribution", "Limitations",
        "Future_Research_Directions", "Extendable_Research_Opportunities",
        "Evidence_Sections", "Notes",
    ],
    "variable_role_matrix.xlsx": [
        "Paper_ID", "Title", "Year", "Theory", "Context", "Variable_Name",
        "Variable_Role", "Construct_Level", "Hypothesized_Path", "Result", "Path_Direction",
        "Measurement_Source", "Analysis_Method", "Model_Structure", "Evidence_Text",
        "Evidence_Section", "Evidence_Source", "Confidence",
    ],
    "qualitative_mechanism_matrix.xlsx": [
        "Paper_ID", "Title", "Year", "Research_Object", "Research_Setting", "Case_Count",
        "Interviewees", "Coding_Method", "First_Order_Concepts", "Second_Order_Themes",
        "Aggregate_Dimensions", "Process_Mechanism", "Boundary_Conditions",
        "Theoretical_Contribution", "Convertible_Quantitative_Variables", "Convertible_Paths",
        "Convertible_Moderators", "Evidence_Text", "Evidence_Section", "Confidence",
    ],
    "theory_gap_matrix.xlsx": [
        "Gap_ID", "Gap_Type", "Gap_Statement", "Supporting_Papers", "Evidence_Texts",
        "Affected_Theory", "Variables_Involved", "Contexts_Involved", "Methodological_Basis",
        "Why_It_Matters", "Topic_Implication", "Confidence",
    ],
}

MARKDOWN_TEMPLATES = {
    "variable_network_summary.md": """# Variable Network Summary

## Corpus Scope and Evidence Limits

## High-Frequency Independent Variables

## High-Frequency Dependent Variables

## High-Frequency Mediators

## High-Frequency Moderators

## Overused or Low-Innovation Variables

## Role-Repositioning Opportunities

## Rare but Theoretically Connectable Combinations

## Antecedent-Mechanism-Outcome-Boundary Models
""",
    "model_candidates.md": """# Model Candidates

For each model: name, management problem, research object/context, antecedent, mechanism, outcome, boundary condition, theoretical rationale, why it is not simple variable拼接, suggested method, data source, publication risk, evidence base, and confidence.
""",
    "topic_cards.md": """# Topic Cards

Generate at least 10 evidence-grounded management empirical topic cards.
""",
    "final_research_story.md": """# Final Research Story

## Top 1

## Top 2

## Top 3

## Next Steps for Introduction and Hypotheses
""",
    "reviewer_evaluation.md": """# Reviewer Evaluation

Evaluate every topic using SSCI management reviewer criteria.
""",
}


def write_workbook(path: Path, columns: list[str]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = path.stem[:31]
    sheet.append(columns)
    sheet.freeze_panes = "A2"
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for index, column in enumerate(columns, start=1):
        sheet.column_dimensions[sheet.cell(row=1, column=index).column_letter].width = min(max(len(column) + 2, 14), 36)
    workbook.save(path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Create empty output templates for the paper topic builder skill.")
    parser.add_argument("--output", default="output", help="Output directory for templates.")
    args = parser.parse_args()

    output_dir = Path(args.output)
    output_dir.mkdir(parents=True, exist_ok=True)

    for filename, columns in SCHEMAS.items():
        write_workbook(output_dir / filename, columns)

    for filename, content in MARKDOWN_TEMPLATES.items():
        (output_dir / filename).write_text(content, encoding="utf-8")

    print(f"Created {len(SCHEMAS) + len(MARKDOWN_TEMPLATES)} templates in {output_dir}")


if __name__ == "__main__":
    main()
